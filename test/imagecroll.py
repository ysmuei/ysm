import os
import time
import openpyxl
import requests
from bs4 import BeautifulSoup as bs
import pandas as pd
import dload
import base64
from PIL import Image
import imagehash
from PIL import Image, ImageSequence
import natsort
import cv2

#이미지 확장자 알아내기------------------------
from io import BytesIO
import requests, imghdr
def find_extension(url):
    img = BytesIO(requests.get(url).content)
    return imghdr.what(img)
#---------------------------------------------

start_time = time.time()
headers = {"User-Agent" : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36"}

url = "https://v73.danbimovie.vip/"  
sitename = "단비무비"
name = "danbimovie"

hash_img = ""
#res = requests.get(url)
res = requests.get(url, headers=headers)

print("상태코드 : "+ str(res.status_code))
res.raise_for_status()
soup = bs(res.content, "html.parser")

#atages = bs(res.content.decode('utf-8', 'replace'), "html.parser").select("a")
atages = soup.select("a")
print(atages)
print("상태코드 : "+ str(res.status_code))
title = soup.find("title").getText()
# name = title.split(" ")

img_folder = './{}/images'.format(name)

#폴더가 없으면 생성.
if not os.path.isdir(img_folder):
    os.makedirs(img_folder)

src_data=[]
href_data=[]
src_base64_str=""
count = 1

#엑셀 초기값 설정
excel_file = openpyxl.Workbook()
excel_sheet = excel_file.active
excel_sheet.title = 'link'
excel_sheet.append(['Site name','URL','base64 encoding', 'hash','배너주소'])
excel_sheet.column_dimensions['A'].width = 15
excel_sheet.column_dimensions['B'].width = 20
excel_sheet.column_dimensions['C'].width = 50
excel_sheet.column_dimensions['D'].width = 20
excel_sheet.column_dimensions['E'].width = 40

#a태그에서 src와 href속성 추출
for i in atages:
    if 'src=' in str(i) and 'href=' in str(i):
        if 'logo' in str(i) or 'html' in str(i) or 'icon' in str(i):
            continue
        
        src = str(i)
        src = src[src.find('src=')+5:]
        src = src[:src.find('"')]
        if not 'http' in src: 
            if url[-1]=="/":
                src = url[0:-1]+src
            else:
                src = url+src
        x = find_extension(src)#src의 확장자 알아오기.
        if x != 'gif':
            continue
        src_data.append(src)

        href = str(i)
        href = href[href.find('href=')+6:]
        href = href[:href.find('"')]
        if not 'http' in href:
            if url[-1]=="/":
                href = url[0:-1]+href
            else:
                href = url+href
        href_data.append(href)

        excel_sheet.append([sitename, url, src_base64_str, hash_img, href])
        excel_sheet.cell(row=1+count, column=2).hyperlink = url
        excel_sheet.cell(row=1+count, column=5).hyperlink = href
    
        x = find_extension(src)#src의 확장자 알아오기.
        if x=='gif':
            dload.save(src, "{}/{}{}".format(img_folder,name, count)+".gif")
            print("{}{}.gif저장완료!".format(name,count))
         
        else:
            dload.save(src, "{}/{}{}".format(img_folder,name, count)+".png")
            print("{}{}.png저장완료!".format(name,count))
       
        excel_file.save('{}/../{}.xlsx'.format(img_folder,name))
        count+=1


list_images = os.listdir("{}".format(img_folder))
print(list_images)
list_len = len(list_images)
print(list_len)

for i in range(1, list_len+1):
    try:
        with open("{}/{}{}".format(img_folder,name, i)+".gif", 'rb') as img:
            base64_string = base64.b64encode(img.read())
    except:
        with open("{}/{}{}".format(img_folder,name, i)+".png", 'rb') as img:
            base64_string = base64.b64encode(img.read())
    try:
        hash_img = imagehash.average_hash(Image.open("{}/{}{}".format(img_folder,name, i)+".gif"))
    except:
        hash_img = imagehash.average_hash(Image.open("{}/{}{}".format(img_folder,name, i)+".png"))
    print(hash_img)
    excel_sheet.cell(row=1+i, column=3, value="{}".format(base64_string))
    excel_sheet.cell(row=1+i, column=4, value="{}".format(hash_img))
    excel_file.save('{}/../{}.xlsx'.format(img_folder,name))

print(name)
print("크롤링 걸린 시간 : " + str(time.time() - start_time))
excel_file.close()




#전처리코드
dest = "./dest/{}".format(name)
# src_path = "./wfwf"
if not os.path.isdir(dest):
    os.makedirs(dest)
j=0
target_dir = img_folder
file_lists = natsort.natsorted(os.listdir(target_dir))
for file in file_lists:
    with Image.open(target_dir + "\\"+ file) as im:
        j=j+1
        if(".png" in (file[-4:]).lower()):
            name = os.path.join(dest, "{}_{}.png".format(j, idx+1))
            img = cv2.imread(target_dir + "\\"+ file)
            img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            img_resize = cv2.resize(img_gray, None, None, 2,2, cv2.INTER_CUBIC)
            cv2.imwrite(name, img_resize)
        for idx, frame in enumerate(ImageSequence.Iterator(im)):
            # print(img.n_frames)
            name = os.path.join(dest, "{}_{}.png".format(j, idx+1))
            print(name) 
            frame.save(name)
            img = cv2.imread(name)
            img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            img_resize = cv2.resize(img_gray, None, None, 2,2, cv2.INTER_CUBIC)
            #print(f"%s" %name)
            cv2.imwrite(name, img_resize)

#ocr 
from PIL import Image
import easyocr
import os
import natsort
#from konlpy.tag import *
#from konlpy.utils import pprint
from collections import OrderedDict
import argparse
import openpyxl
import csv
from multiprocessing import Process, freeze_support

#hannanum = Hannanum()
#kkma = Kkma()

excel = openpyxl.Workbook()
excel_ws = excel.active
excel_ws.title = 'keyword'

# dest = f"C:\\Users\\ParkSiHyeon\\Desktop\\caps\\dst5\\marumaru"

file_lists = []
for(root, directorys, files) in os.walk(dest):
   for file in files:
      file_path = os.path.join(file)
      file_lists.append(file_path)

file_lists = natsort.natsorted(file_lists)

criteria = ['첫충', '매충', '페이백', '카지노', 
         '라이브카지노', '호텔카지노', '코드', '놀이터', 
         '돌발', '비아그라', '미니게임', '룰렛', '사다리' 
         '신규첫충', '무한매충', '무한첫충', '신규', '베팅','단폴', '두폴'
         ]

def fn_infer():
   a= str()
   tmp1= list()
   target = "TWICE"
   for file in file_lists:
      tmp = list()
      if(target != file.split("_")[0]):
         a = ' '.join(OrderedDict.fromkeys(tmp1))
         kxl = a.split()
         length = len(kxl)
         print(a)
         
         if target=="TWICE":
            target = file.split("_")[0] 
         
         excel_ws.cell(row=int(target),column=1,value=a) #한번에 때려박는거
         for i in range(0, length):
            excel_ws.cell(row=int(target), column=2+i, value=kxl[i]) #한칸씩 넣는거
         excel.save('test_danbi.xlsx')
         tmp1= list()
         target = file.split("_")[0] 
         print(target)
         
      
      #result = demo.demo(opts)
      reader = easyocr.Reader(['ko', 'en'], gpu=True)
      result = reader.readtext(dest + "/"+ file, detail=0)

      print(result)
      #result = ''.join(result)
      #NLPBF = kkma.nouns(result)
      #print("After NLP : ", NLPBF)

      #문자열 비교 시작
      for i in criteria:
         for j in result:
            if i in j:
               tmp.append(i)
               tmp1 = tmp
         

if __name__ == '__main__':
   freeze_support()
   fn_infer()